import tkinter as tk
from tkinter import ttk,filedialog,font
import csv
import threading
import datetime
from netmiko import ConnectHandler
import time
import difflib
import os
from openpyxl import Workbook, load_workbook

def push_config_gui(push_header,push_frame):
  command_list = []
  
  def load_cmd():
    filename = filedialog.askopenfilename(title='Select capture commmand file (.txt)', filetypes=(("TXT files", "*.txt"), ("All files", "*.*")))
    if filename:
      cmd_file_var.set(f"{os.path.basename(filename)}")
      with open(filename, 'r') as file:
        command_list.clear()
        for line in file:
          command_list.append(line.strip())
        
        print(command_list)
  
  def load_csv():
    filename = filedialog.askopenfilename(title="Select CSV File", filetypes=(("CSV files", "*.csv"), ("All files", "*.*")))
    if filename:
      try:
        with open(filename, newline='') as file:
          reader = csv.DictReader(file)
          data = [row for row in reader]
        return data
      except FileNotFoundError:
        return None
  
  def display_push():
    # Clear existing data in the Treeview
    for item in push_tree.get_children():
      push_tree.delete(item)

    data = load_csv()
    if data:
      # Create header row
      for i, item in enumerate(push_header):
        push_tree.heading("#" + str(i+1), text=item)
        push_tree.column("#" + str(i+1), stretch=True, width=100)

      # Start SSH process for each bulk
      for index,push in enumerate(data):
        push['no'] = index + 1
        row_data = [push['no'], push['ip'], '','','','','','','']
        item_id = push_tree.insert("", "end", values=row_data)
        # Start SSH process for each bulk
        threading.Thread(target=push_config, args=(push, item_id)).start()
  
  def push_config(push, item_id):
    current_datetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    error_filename = f"errorlogs_{current_datetime}.xlsx"
    number = push['no']
    vendor = push['vendor']
    ip = push['ip']
    user = push['user']
    filepath = push['filepath']
    file = push['file']
    password = push['password']
    protocol = push['protocol']
    server = push['tftp_ip']
    tftp_address = f'{server}/{filepath}/{file}'
    capture_log_path = capturelog_entry.get()
    log_path = error_path_entry.get()
    condition = select_condition_var.get()
    bool_reload = reload_var.get()
    minutes = reload_entry.get()
    
    if vendor == 'cisco' and protocol == 'telnet':
      device = {
        'device_type': 'cisco_ios_telnet',
        'host': ip,
        'username': user,
        'password': password,
        'port': 23,
      }
    elif vendor == 'cisco' and protocol == 'ssh':
      device = {
        'device_type': 'cisco_ios',
        'host': ip,
        'username': user,
        'password': password,
        'port': 22,
      }
    
    def check_connection(device):
      try:
        connect = ConnectHandler(**device)
        return connect, None
      except Exception as e:
        return None, str(e)

    net_connect, error = check_connection(device)
    if error:
      tk.messagebox.showinfo('Error on process no. ' + str(number), error)
      push_tree.set(item_id, "#3", "Fail")
      push_tree.update_idletasks()
      return
    else:
      net_connect.find_prompt()
      push_tree.set(item_id, '#3', "Connected")
      push_tree.update_idletasks()
      
      if condition == 'existing':
        if bool_reload == 'yes' and minutes:
          out = net_connect.send_command('reload in ' + minutes,read_timeout=100, expect_string=' [confirm]')
          print(out)
          net_connect.send_command_timing('\n',read_timeout=200)
          push_tree.set(item_id, '#4', minutes + ' minutes')
          push_tree.update_idletasks()
          
          time.sleep(2)
        elif bool_reload == 'yes' and minutes != '':
          tk.messagebox.showinfo('Error', "Please input the number of minutes for reload")
          return
        else:
          push_tree.set(item_id, '#4', 'No reload')
          push_tree.update_idletasks()
        
        errors = ''
        push_tree.set(item_id, '#5', 'Retrieving Current status...')
        push_tree.update_idletasks()
        currentint = net_connect.send_command('sh ip int br', read_timeout=300)
        print(currentint)
        time.sleep(0.5)
        push_tree.set(item_id, '#5', 'Retrieved!')
        push_tree.update_idletasks()
        date = datetime.datetime.now().strftime("%Y-%m-%d")
        hostname = net_connect.send_command('sh run | include hostname').split()[1]
        filename = f'{hostname}_{date}.cfg'
        push_tree.set(item_id, '#6', 'Backing Up...')
        push_tree.update_idletasks()
        
        try:
          output = net_connect.send_command_timing('copy running-config flash:' + filename, read_timeout=300)
          print(output)
          output = net_connect.send_command_timing('\n', read_timeout=300)
          print(output)
          
          net_connect.send_command('\n')
          net_connect.send_command('\n')
          push_tree.set(item_id, '#6', 'Backup Success')
          push_tree.update_idletasks()
          
        except Exception as e:
          tk.messagebox.showinfo('Process No. ' + str(number), "Fail Backing Up")
          push_tree.set(item_id, '#6', 'Backup Failed')
          push_tree.update_idletasks()
        
        push_tree.set(item_id, "#7", "Copying..")
        push_tree.update_idletasks()
        net_connect.send_command_timing('copy tftp://' + tftp_address + ' flash:', read_timeout=300)
        output = net_connect.send_command('\n', read_timeout=100, expect_string='Accessing tftp://')
        print(output)
        output = net_connect.send_command_timing('dir flash:')
        
        if file in output:
          push_tree.set(item_id, "#7", "Copy Success")
          push_tree.update_idletasks()
        elif file not in output:
          push_tree.set(item_id, "#7", "Copy Fail")
          push_tree.update_idletasks()
          return
        
        output = net_connect.send_command_timing('copy flash:' + file + ' running-config',read_timeout=300)
        print(output)
        push_result = net_connect.send_command('\n', read_timeout=300, expect_string=r' bytes/sec\)')
        print(push_result)
        
        push_tree.set(item_id, "#8", "Pushing Config...")
        push_tree.update_idletasks()
        
        if "Invalid" in push_result:
          error_lines = []
          lines = push_result.splitlines()
          for line in lines:
              if 'Invalid' in line or '^' in line or 'bytes' in line:
                  continue  # Skip lines containing these error messages
              error_lines.append(line) 

          errors = "\n".join(error_lines)
          
          push_tree.set(item_id, "#8", "Write memory...")
          push_tree.update_idletasks()
          net_connect.send_command('wr mem')
          time.sleep(1)
          
          if bool_reload == 'yes':
            push_tree.set(item_id, "#8", "Invalid cmd, cancel reload..")
            push_tree.update_idletasks()
            net_connect.send_command('reload cancel')
            time.sleep(3)
            push_tree.set(item_id, "#8", "Invalid cmd, reload cancelled")
            push_tree.update_idletasks()
          else:
            push_tree.set(item_id, "#8", "Invalid cmd found")
            push_tree.update_idletasks() 

        else:
          errors = 'None'
          net_connect.send_command('\n')
          net_connect.send_command('\n')
          push_tree.set(item_id, "#8", "Write memory...")
          push_tree.update_idletasks()
          net_connect.send_command('wr mem')
          time.sleep(1)
          
          if bool_reload == 'yes':
            push_tree.set(item_id, "#8", "Success, cancel reload..")
            push_tree.update_idletasks()
            net_connect.send_command('reload cancel')
            time.sleep(3)
            push_tree.set(item_id, "#8", "Success, reload cancelled")
            push_tree.update_idletasks()
          else:
            push_tree.set(item_id, "#8", "Success")
            push_tree.update_idletasks()
        
        afterint = net_connect.send_command('sh ip int br', read_timeout=300)
        afterint_lines = afterint.strip().split('\n')
        currentint_lines = currentint.strip().split('\n')
        intdiff = difflib.unified_diff(currentint_lines, afterint_lines)
        output_string = ""
        # Check if there are differences
        if any(intdiff):
          added_lines =[]
          removed_lines = []
          push_tree.set(item_id, "#9", "Changes found")
          push_tree.update_idletasks()
          # Classify lines as added (+) or removed (-)
          for line in intdiff:
            if line.startswith('+'):
              added_lines.append(line[1:])  
            elif line.startswith('-'):
              removed_lines.append(line[1:])  

          output_string += "Added lines:\n"
          for line in added_lines:
              output_string += f"+ {line}\n"

          output_string += "\n"

          output_string += "Removed lines:\n"
          for line in removed_lines:
              output_string += f"- {line}\n"
           
        else:
          # If no differences, display confirmation message
          push_tree.set(item_id, "#9", "Verified!")
          push_tree.update_idletasks()
          output_string = 'No changes'
        
        if capt_var.get() == 'yes':
          push_tree.set(item_id, "#10", "Capturing...")
          push_tree.update_idletasks()
            
          hostname = net_connect.send_command('sh run | include hostname').split()[1]
          date_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
          logfilename = f'{hostname}_{date_time}.txt'
          net_connect.send_command('term length 0')
          hostname = net_connect.send_command('sh run | include hostname').split()[1]
              
          output = ''
          for cmd in command_list:
            print(f"Sending command: {cmd}")  # Debugging print
            cmd_output = net_connect.send_command(cmd,read_timeout=600)
            print(f"Command: {cmd}\nOutput: {cmd_output}")  # Debugging print
            output += f'{hostname}#{cmd}\n{cmd_output}\n\n'
            time.sleep(0.5)
        
          with open(os.path.join(capture_log_path,logfilename),'w') as log_file:
            log_file.write(output)
          
          push_tree.set(item_id, "#10", "Done Capturing")
          push_tree.update_idletasks()
        
        else:
          push_tree.set(item_id, "#10", "Not Capturing")
          push_tree.update_idletasks()

        if not os.path.exists(error_file):
          # Create a new Excel workbook
          wb = Workbook()
          ws = wb.active
          ws.append(['Hostname',"IP Address", 'Errors', 'Changes in interface'])

          # Save the workbook initially
          wb.save(error_file)
        else:
          # Load existing workbook
          wb = load_workbook(error_file)
          ws = wb.active

        ws.append([hostname, ip, errors, output_string])
        # Save workbook
        wb.save(error_file)
        
        if 'Invalid' in push_result and not 'No changes' in output_string:
          tk.messagebox.showinfo('Error on process ' + str(number), f'Invalid command & changes in interface status found. Output saved to {error_file}')
        elif 'Invalid' in push_result and 'No changes' in output_string:
          tk.messagebox.showinfo('Error on process ' + str(number), f'Invalid command found. Output saved to {error_file}')
        elif not 'Invalid' in push_result and not 'No changes' in output_string:
          tk.messagebox.showinfo('Error on process ' + str(number), f'Changes found in interface status. Output saved to {error_file}')

      elif condition == 'staging':
        if bool_reload == 'yes' and minutes:
          out = net_connect.send_command('reload in ' + minutes, read_timeout=100, expect_string=' [confirm]')
          print(out)
          net_connect.send_command_timing('\n',read_timeout=200)
          push_tree.set(item_id, '#4', minutes + ' minutes')
          push_tree.update_idletasks()
          time.sleep(2)
        elif bool_reload == 'yes' and minutes != '':
          tk.messagebox.showinfo('Error', "Please input the number of minutes for reload")
          return
        else:
          push_tree.set(item_id, '#4', 'No reload')
          push_tree.update_idletasks()

        errors = ''
        push_tree.set(item_id, '#5', 'Retrieving Current status...')
        push_tree.update_idletasks()
        currentint = net_connect.send_command('sh ip int br', read_timeout=300)
        print(currentint)
        time.sleep(0.5)
        push_tree.set(item_id, '#5', 'Retrieved!')
        push_tree.update_idletasks()
        
        push_tree.set(item_id, '#6', 'Staging: No Backup')
        push_tree.update_idletasks()
        
        net_connect.send_command_timing('copy tftp://' + tftp_address + ' flash:', read_timeout=300)
        output = net_connect.send_command('\n', read_timeout=100, expect_string='Accessing tftp://')
        print(output)
        push_tree.set(item_id, "#7", "Copying..")
        push_tree.update_idletasks()
        
        output = net_connect.send_command_timing('dir flash:')
        
        if file in output:
          push_tree.set(item_id, "#7", "Copy Success")
          push_tree.update_idletasks()
        else:
          push_tree.set(item_id, "#7", "Copy Fail")
          push_tree.update_idletasks()
          return
        
        output = net_connect.send_command_timing('copy flash:' + file + ' running-config',read_timeout=300)
        print(output)
        push_result = net_connect.send_command('\n', read_timeout=300, expect_string=r' bytes/sec\)')
        print(push_result)
        
        push_tree.set(item_id, "#8", "Pushing Config...")
        push_tree.update_idletasks()
        
        if "Invalid" in push_result:
          error_lines = []
          lines = push_result.splitlines()
          for line in lines:
              if 'Invalid' in line or '^' in line or 'bytes' in line:
                continue  # Skip lines containing these error messages
              error_lines.append(line)  

          errors = "\n".join(error_lines)
          
          push_tree.set(item_id, "#8", "Write memory...")
          push_tree.update_idletasks()
          net_connect.send_command('wr mem')
          time.sleep(1)
          if bool_reload == 'yes':
            push_tree.set(item_id, "#8", "Invalid cmd, cancel reload..")
            push_tree.update_idletasks()
            net_connect.send_command('reload cancel')
            time.sleep(3)
            push_tree.set(item_id, "#8", "Invalid cmd, reload cancelled")
            push_tree.update_idletasks()
          else:
            push_tree.set(item_id, "#8", "Invalid cmd found")
            push_tree.update_idletasks() 

        else:
          errors = 'None'
          net_connect.send_command('\n')
          net_connect.send_command('\n')
          push_tree.set(item_id, "#8", "Write memory...")
          push_tree.update_idletasks()
          net_connect.send_command('wr mem')
          time.sleep(1)
          
          if bool_reload == 'yes':
            push_tree.set(item_id, "#8", "Success, cancel reload..")
            push_tree.update_idletasks()
            net_connect.send_command('reload cancel')
            time.sleep(3)
            push_tree.set(item_id, "#8", "Success, reload cancelled")
            push_tree.update_idletasks()
          else:
            push_tree.set(item_id, "#8", "Success")
            push_tree.update_idletasks()
          
        
        afterint = net_connect.send_command('sh ip int br', read_timeout=300)
        afterint_lines = afterint.strip().split('\n')
        currentint_lines = currentint.strip().split('\n')
        intdiff = difflib.unified_diff(currentint_lines, afterint_lines)
        output_string = ""
        # Check if there are differences
        if any(intdiff):
          error_file = os.path.join(log_path, error_filename)
          added_lines =[]
          removed_lines = []
          push_tree.set(item_id, "#9", "Changes found")
          push_tree.update_idletasks()
          # Classify lines as added (+) or removed (-)
          for line in intdiff:
            if line.startswith('+'):
              added_lines.append(line[1:])  
            elif line.startswith('-'):
              removed_lines.append(line[1:])  

          output_string += "Added lines:\n"
          for line in added_lines:
              output_string += f"+ {line}\n"

          output_string += "\n"

          output_string += "Removed lines:\n"
          for line in removed_lines:
              output_string += f"- {line}\n"
          
          
        else:
          # If no differences, display confirmation message
          push_tree.set(item_id, "#9", "Verified!")
          push_tree.update_idletasks()
          output_string = 'No changes'
        
        if capt_var.get() == 'yes':
          push_tree.set(item_id, "#10", "Capturing...")
          push_tree.update_idletasks()
            
          hostname = net_connect.send_command('sh run | include hostname').split()[1]
          date_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
          logfilename = f'{hostname}_{date_time}.txt'
          net_connect.send_command('term length 0')
          hostname = net_connect.send_command('sh run | include hostname').split()[1]
              
          output = ''
          for cmd in command_list:
            print(f"Sending command: {cmd}")  # Debugging print
            cmd_output = net_connect.send_command(cmd,read_timeout=600)
            print(f"Command: {cmd}\nOutput: {cmd_output}")  # Debugging print
            output += f'{hostname}#{cmd}\n{cmd_output}\n\n'
            time.sleep(0.5)
        
          with open(os.path.join(capture_log_path,logfilename),'w') as log_file:
            log_file.write(output)
          
          push_tree.set(item_id, "#10", "Done Capturing")
          push_tree.update_idletasks()
        
        else:
          push_tree.set(item_id, "#10", "Not Capturing")
          push_tree.update_idletasks()

        if not os.path.exists(error_file):
          # Create a new Excel workbook
          wb = Workbook()
          ws = wb.active
          ws.append(["IP Address", 'Errors', 'Changes in interface'])

          # Save the workbook initially
          wb.save(error_file)
        else:
          # Load existing workbook
          wb = load_workbook(error_file)
          ws = wb.active

        ws.append([ip, errors, output_string])
        # Save workbook
        wb.save(error_file)
        
        if 'Invalid' in push_result and not 'No changes' in output_string:
          tk.messagebox.showinfo('Error on process ' + str(number), f'Invalid command & changes in interface status found. Output saved to {error_file}')
        elif 'Invalid' in push_result and 'No changes' in output_string:
          tk.messagebox.showinfo('Error on process ' + str(number), f'Invalid command found. Output saved to {error_file}')
        elif not 'Invalid' in push_result and not 'No changes' in output_string:
          tk.messagebox.showinfo('Error on process ' + str(number), f'Changes found in interface status. Output saved to {error_file}')

        
  style = ttk.Style()
  style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])  
  style.layout("Custom.Treeview.Heading", style.layout("Treeview.Heading"))

  style.configure("Custom.Treeview", background="#D3D3D3")
  style.map("Custom.Treeview", foreground=[('selected', 'black')])
  
  bold_font = font.Font(family="TkDefaultFont", weight="bold",size=10)
  
  info_label = ttk.Label(push_frame, text='CSV Format = vendor, ip, user, password, protocol (telnet/ssh), tftp_ip, filepath, file', font=bold_font)
  info_label.grid(row=1, column=0, padx=(0,10), pady=5,columnspan=3)
  
  capturelog_label = ttk.Label(push_frame, text='Path to Store Capture Logs:')
  capturelog_label.grid(row=2, column=0,columnspan=3)
  capturelog_entry = ttk.Entry(push_frame, width=23)
  capturelog_entry.grid(row=3, column=0,columnspan=3)
  
  # Select condition
  select_condition_label = ttk.Label(push_frame, text='Device Condition')
  select_condition_label.grid(row=4, column=0,padx=(250,0))
  select_condition_var = tk.StringVar(value='staging')
  select_condition_dropdown = ttk.Combobox(push_frame, textvariable=select_condition_var, values=['existing', 'staging'])
  select_condition_dropdown.grid(row=5, column=0,padx=(250,0))

  error_path_label = ttk.Label(push_frame, text='Push Config Log Path:')
  error_path_label.grid(row=6, column=0,padx=(250,0))
  error_path_entry = ttk.Entry(push_frame, width=23)
  error_path_entry.grid(row=7, column=0,padx=(250,0),pady=(0,10))
  
  # Use capture dropdown
  capt_label = ttk.Label(push_frame, text='Use capture config:')
  capt_label.grid(row=4, column=1)
  capt_var = tk.StringVar(value='no')
  capt_dropdown = ttk.Combobox(push_frame, textvariable=capt_var, values=['yes', 'no'])
  capt_dropdown.grid(row=5, column=1)

  cmd_file_var = tk.StringVar(value='Import')
  # capture file select
  capt_file_label = ttk.Label(push_frame, text='Select capture file:')
  capt_file_label.grid(row=6, column=1)
  open_txt_button = ttk.Button(push_frame, textvariable=cmd_file_var, command=load_cmd,state='disabled')
  open_txt_button.grid(row=7, column=1,pady=(0,10))

  # Enable/disable capture based on dropdown selection
  def on_capt_change(event):
    if capt_var.get() == 'yes':
      open_txt_button.config(state='normal')
    else:
      open_txt_button.config(state='disabled')

  capt_dropdown.bind('<<ComboboxSelected>>', on_capt_change)
  # Use reload dropdown
  reload_label = ttk.Label(push_frame, text='Use Reload:')
  reload_label.grid(row=4, column=2,columnspan=1,padx=(0,250))
  reload_var = tk.StringVar(value='no')
  reload_dropdown = ttk.Combobox(push_frame, textvariable=reload_var, values=['yes', 'no'])
  reload_dropdown.grid(row=5, column=2,columnspan=1,padx=(0,250))

  # Reload time entry
  reload_time_label = ttk.Label(push_frame, text='Reload Time (minutes):')
  reload_time_label.grid(row=6, column=2,columnspan=1,padx=(0,250))
  reload_entry = ttk.Entry(push_frame, width=23, state='disabled')
  reload_entry.grid(row=7, column=2,columnspan=1,padx=(0,250),pady=(0,10))

  # Enable/disable reload time entry based on dropdown selection
  def on_reload_change(event):
    if reload_var.get() == 'yes':
      reload_entry.config(state='normal')
    else:
      reload_entry.config(state='disabled')

  reload_dropdown.bind('<<ComboboxSelected>>', on_reload_change)

  # Button to open file explorer
  open_button = ttk.Button(push_frame, text="Import CSV & Start", command=display_push)
  open_button.grid(row=10, column=0, padx=10,pady=5,columnspan=3)

  push_tree = ttk.Treeview(push_frame, columns=[f"#{i}" for i in range(1, len(push_header) + 1)], show="headings", height=25, style="Custom.Treeview")
  push_tree.grid(row=11, column=0, sticky="nsew",columnspan=3)
  for i, col in enumerate(push_header):
    push_tree.heading(f"#{i+1}", text=col)
    push_tree.column(f"#{i+1}", width=100, stretch=tk.YES)

  # Add scrollbar to the Treeview
  scrollbar = ttk.Scrollbar(push_frame, orient="vertical", command=push_tree.yview)
  push_tree.configure(yscroll=scrollbar.set)
  scrollbar.grid(row=11, column=4, sticky="ns")

  push_frame.grid_rowconfigure(3, weight=1)
  push_frame.grid_columnconfigure(1, weight=1)
  push_frame.pack()
  